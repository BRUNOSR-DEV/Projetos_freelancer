from utils import helper
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar as calendario


# função salvar valor
valores = []
def salvar_valor():
    valor_dia = float(entry_valor.get())
    valores.append(valor_dia)
    total_valores = sum(valores)
    entry_valor.delete(0,tk.END)

    label_status.config(text="Valor salvo com sucesso!", foreground="green")
    label_total.config(text=f"Total economizado: R${total_valores:.2f}")

    #ADicionando um novo valor em uma linha da planilha
    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1, value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha,column=2, value= valor_dia)

janela = tk.Tk()
janela.title('App de Poupança Pessoal') # Título que ficará no topo da janela
janela.geometry('855x720')  # Dimenção da jenela criada
janela.configure(bg= "#252525")


style = ttk.Style()
style.theme_use("clam")
style.configure('Tlabel', background = "#252525", foreground="#FFFFFF", font=("Arial", 12))
style.configure('TEntry', fieldbackground = "#FFFFFF", font=("Arial", 12))
style.configure('TButton', background = "#4caf50", foreground="#FFFFFF", font=("Arial", 12))

label_instrucao = ttk.Label(janela, text= "Insira o valor diário: ")
label_status = ttk.Label(janela, text= "mmmmmmmmm",foreground="red")
label_total = ttk.Label(janela, text= "kkkkkkk",font=("Arial", 14, "bold"))
entry_valor = ttk.Entry(janela)
button_salvar = ttk.Button(janela, text="Salvar", command= salvar_valor)

#posicionando elementos

label_instrucao.pack(pady=10)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack()
label_total.pack(pady= 10)


# carregamento da planilha existente ou criação de uma nova

try:
    workbook =  load_workbook("valores_diarios.xlsx")
except FileNotFoundError:
    workbook = Workbook()

# Selecionando a primeira planilha  
sheet =  workbook.active

# Verificando se a planilha já possui valores salvos

if sheet.max_row == 0:
    sheet.cell(row=1, column=1, value='Data')
    sheet.cell(row=1, column=2, value='Diário')

# Obtem a lista de valores já salvos
    
valores =  [cell.value for cell in sheet['B'][1:]]

# Exibe o total economizado
label_total.config(text=f'Total economizado: R${sum(valores):.2f}')








janela.mainloop()



# Salva planilha com valores atualizados
workbook.save('valores_diarios.xlsx')
